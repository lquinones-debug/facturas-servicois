"""app.py — Control de Facturas de Servicios (app web multiusuario).

Streamlit + Google Sheets. Base de datos COMPARTIDA: lo que carga una persona y
lo que marca "Pagos" lo ven todos en tiempo real (no es localStorage por navegador).

Secciones:
  • Cargar factura  — formulario de alta
  • Pagos           — buscar comprobante y tildar pagado
  • Consultar       — pendientes / pagadas / histórico por períodos
  • Proveedores     — alta de proveedores nuevos (para que aparezcan en Cargar)

Correr local:   streamlit run tools/facturas_app/app.py
"""

from __future__ import annotations

import re
from datetime import date, timedelta

import pandas as pd
import streamlit as st

import sheet_db as db
import drive_db as drive
import factura_logic as fl

st.set_page_config(page_title="Control de Facturas - Servicios", page_icon="🧾", layout="wide")


# ----- Acceso (clave única compartida) ---------------------------------------

def _clave_app() -> str:
    """Clave compartida: st.secrets['app_password'] (nube) o FACTURAS_APP_PASSWORD;
    por defecto 00000000. Se usa para entrar y para confirmar borrados."""
    import os
    try:
        clave = st.secrets.get("app_password")
    except Exception:
        clave = None
    return str(clave or os.getenv("FACTURAS_APP_PASSWORD", "00000000"))


def _check_password() -> bool:
    """Portón de contraseña (clave compartida del equipo)."""
    clave = _clave_app()

    if st.session_state.get("auth_ok"):
        return True

    st.title("🔒 Control de Facturas de Servicios")
    st.caption("Ingresá la clave de acceso del equipo.")
    pwd = st.text_input("Contraseña", type="password", key="login_pwd")
    if st.button("Entrar", type="primary"):
        if pwd == str(clave):
            st.session_state["auth_ok"] = True
            st.rerun()
        else:
            st.error("Contraseña incorrecta.")
    return False


if not _check_password():
    st.stop()


# ----- Helpers ----------------------------------------------------------------

def _refrescar():
    db.invalidar_cache()
    st.cache_data.clear()


@st.cache_data(ttl=db.CACHE_TTL, show_spinner=False)
def cargar_facturas():
    return db.listar_facturas(force_refresh=True)


@st.cache_data(ttl=db.CACHE_TTL, show_spinner=False)
def cargar_proveedores():
    return db.listar_proveedores(force_refresh=True)


@st.cache_data(ttl=db.CACHE_TTL, show_spinner=False)
def cargar_descargas():
    return db.claves_descargadas(force_refresh=True)


def fmt_money(v):
    if v in (None, ""):
        return ""
    try:
        return f"$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return str(v)


def parse_monto_ar(s):
    """Parsea un monto en formato argentino: '.' = miles, ',' = decimales.

    '27.000,00' -> 27000.0 ; '27000' -> 27000.0 ; vacío/inválido -> None.
    """
    s = (s or "").strip().replace(" ", "").replace("$", "")
    if not s:
        return None
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def formatear_monto_ar(s):
    """Formatea un monto al estilo argentino con separador de miles.

    '56000' -> '56.000' ; '56000,5' -> '56.000,50'. None si no parsea.
    """
    v = parse_monto_ar(s)
    if v is None:
        return None
    if v == int(v):
        return f"{int(v):,}".replace(",", ".")
    return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _autoformatear_monto():
    """Callback on_change del campo monto: lo reformatea con miles al salir."""
    f = formatear_monto_ar(st.session_state.get("cf_monto", ""))
    if f is not None:
        st.session_state["cf_monto"] = f


def periodo_auto(primer_vto):
    """Período = mes anterior al 1er vencimiento, formato 'MM-AA' (ej. 05-26)."""
    prev = primer_vto.replace(day=1) - timedelta(days=1)
    return prev.strftime("%m-%y")


def formatear_comprobante(s):
    """Normaliza a 'XXXXX - XXXXXXXX' (5 + 8 dígitos, rellenando ceros a la izquierda).

    '67 - 8878' -> '00067 - 00008878'. Devuelve None si no hay dos grupos de dígitos.
    """
    grupos = [g for g in re.split(r"\D+", (s or "").strip()) if g]
    if len(grupos) != 2:
        return None
    izq, der = grupos
    return f"{int(izq):05d} - {int(der):08d}"


def _autoformatear_comp():
    """Callback on_change del campo comprobante: lo reformatea al salir del campo."""
    f = formatear_comprobante(st.session_state.get("cf_comp", ""))
    if f:
        st.session_state["cf_comp"] = f


def estado_venc(f):
    """(codigo, etiqueta, emoji) de estado de vencimiento de una factura dict del sheet."""
    vto = fl.parse_fecha(f.get("primer_vto"))
    cod, et = fl.estado_factura(vto, date.today())
    return cod, et, fl.EMOJI_ESTADO.get(cod, "⚪")


def periodo_par(f):
    """Devuelve (orden 'YYYY-MM', etiqueta 'MM-YY') del período de una factura.

    Usa el campo `periodo` (MM-YY) si está; si no, lo deriva del 1er vencimiento.
    """
    p = (f.get("periodo") or "").strip()
    m = re.match(r"(\d{1,2})-(\d{2})$", p)
    if m:
        return f"20{m.group(2)}-{int(m.group(1)):02d}", f"{int(m.group(1)):02d}-{m.group(2)}"
    vto = fl.parse_fecha(f.get("primer_vto"))
    if vto:
        return vto.strftime("%Y-%m"), vto.strftime("%m-%y")
    return "0000-00", "s/f"


def aplicar_filtros(facturas):
    """Renderiza el panel 🔎 Segmentar / filtrar y devuelve (facturas_filtradas, hay_filtro).

    Compartido por 📊 Consultar y 📈 Dashboard (mismas keys flt_*, así el filtro
    se mantiene al cambiar de sección)."""
    with st.expander("🔎 Segmentar / filtrar", expanded=False):
        f_provs = sorted({f.get("proveedor", "") for f in facturas if f.get("proveedor")})
        f_rubros = sorted({f.get("rubro", "") for f in facturas if f.get("rubro")})
        f_periodos = sorted({f.get("periodo", "") for f in facturas if f.get("periodo")}, reverse=True)
        fc1, fc2, fc3 = st.columns(3)
        sel_prov = fc1.multiselect("Proveedor", f_provs, key="flt_prov")
        sel_rubro = fc2.multiselect("Rubro", f_rubros, key="flt_rubro")
        sel_periodo = fc3.multiselect("Período", f_periodos, key="flt_periodo")
        gc1, gc2 = st.columns([3, 2])
        sel_texto = gc1.text_input(
            "Buscar texto (cuenta, nº cliente, comprobante, nota)", key="flt_texto")
        sel_estado = gc2.radio("Estado de pago", ["Todas", "Pendientes", "Pagadas"],
                               horizontal=True, key="flt_estado")
        if st.button("🧹 Limpiar filtros"):
            for kk in ("flt_prov", "flt_rubro", "flt_periodo", "flt_texto", "flt_estado"):
                st.session_state.pop(kk, None)
            st.rerun()

    def _pasa(f):
        if sel_prov and f.get("proveedor") not in sel_prov:
            return False
        if sel_rubro and f.get("rubro") not in sel_rubro:
            return False
        if sel_periodo and f.get("periodo") not in sel_periodo:
            return False
        if sel_estado == "Pendientes" and f.get("estado_pago") == db.ESTADO_PAGADA:
            return False
        if sel_estado == "Pagadas" and f.get("estado_pago") != db.ESTADO_PAGADA:
            return False
        if sel_texto:
            q = sel_texto.lower()
            campos = [f.get(k, "") for k in
                      ("proveedor", "cuenta", "nro_cliente", "comprobante", "nota", "periodo")]
            if not any(q in str(v).lower() for v in campos):
                return False
        return True

    facturas_f = [f for f in facturas if _pasa(f)]
    hay = bool(sel_prov or sel_rubro or sel_periodo or sel_texto or sel_estado != "Todas")
    if hay:
        st.info(f"🔎 Filtro activo: **{len(facturas_f)}** de {len(facturas)} facturas.")
    return facturas_f, hay


# Rubros (categoría del gasto). Lista fija + "Otros" (texto libre). Editable.
RUBROS = ["Agua", "Alquiler", "Expensas", "Gas", "Internet", "Luz",
          "Seguridad", "Tasa Municipal", "Telefonía", "Otros"]
RUBRO_PLACEHOLDER = "(sin asignar)"
RUBRO_OPCIONES = [RUBRO_PLACEHOLDER] + RUBROS


def _input_rubro(key_prefix: str, actual: str = "") -> str:
    """Menú desplegable de rubro (+ texto libre si es 'Otros'). Devuelve el rubro final."""
    if actual in RUBROS:
        idx = RUBRO_OPCIONES.index(actual)
    elif actual:
        idx = RUBRO_OPCIONES.index("Otros")
    else:
        idx = 0
    sel = st.selectbox("Rubro", RUBRO_OPCIONES, index=idx, key=f"{key_prefix}_rubro")
    if sel == "Otros":
        custom = st.text_input("Especificá el rubro",
                               value=(actual if actual not in RUBROS else ""),
                               key=f"{key_prefix}_rubro_otro")
        return custom.strip()
    if sel == RUBRO_PLACEHOLDER:
        return ""
    return sel


def fmt_monto_ar_num(v):
    """Float -> string argentino sin '$' (para prellenar el campo monto al editar)."""
    if v in (None, ""):
        return ""
    return fmt_money(v).replace("$", "").strip()


# ----- Exportar a Excel (.xlsx) con tabla filtrable / segmentable --------------

# Columnas legibles del export (orden = como se ven en el Excel)
EXPORT_COLS = [
    "Estado pago", "Estado venc.", "Proveedor", "Cuenta", "Nº Cliente", "Rubro",
    "Período", "Emisión", "1er Vto", "2do Vto", "Monto", "Comprobante",
    "Pagada el", "Pagó", "Origen", "Nota",
]


def df_export(facturas_list):
    """DataFrame legible (columnas en español) para mostrar/segmentar/exportar."""
    filas = []
    for f in facturas_list:
        _, et, _ = estado_venc(f)
        filas.append({
            "Estado pago": f.get("estado_pago", db.ESTADO_PENDIENTE),
            "Estado venc.": et,
            "Proveedor": f.get("proveedor", ""),
            "Cuenta": f.get("cuenta", ""),
            "Nº Cliente": f.get("nro_cliente", ""),
            "Rubro": f.get("rubro", ""),
            "Período": f.get("periodo", ""),
            "Emisión": f.get("fecha_emision", ""),
            "1er Vto": f.get("primer_vto", ""),
            "2do Vto": f.get("segundo_vto", ""),
            "Monto": f.get("monto_num") or 0.0,
            "Comprobante": f.get("comprobante", ""),
            "Pagada el": f.get("fecha_pago", ""),
            "Pagó": f.get("pagado_por", ""),
            "Origen": f.get("origen", ""),
            "Nota": f.get("nota", ""),
        })
    return pd.DataFrame(filas, columns=EXPORT_COLS)


def exportar_excel(facturas_list):
    """Genera un .xlsx de la base como **Tabla de Excel** (autofiltro + listo para
    agregar Segmentación de datos: en Excel → pestaña Insertar → Segmentación).
    Devuelve los bytes del archivo."""
    import io
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    df = df_export(facturas_list)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xl:
        df.to_excel(xl, index=False, sheet_name="Facturas")
        ws = xl.sheets["Facturas"]

        nfilas = len(df)
        ncols = len(df.columns)
        ref = f"A1:{get_column_letter(ncols)}{nfilas + 1}"

        # Tabla de Excel: da el autofiltro y habilita Segmentación de datos (slicers)
        if nfilas >= 1:
            tabla = Table(displayName="Facturas", ref=ref)
            tabla.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
            ws.add_table(tabla)

        # Encabezado en negrita + congelar la primera fila
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="305496")
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"

        # Ancho de columnas según contenido
        for i, col in enumerate(df.columns, start=1):
            largos = [len(str(col))] + [len(str(v)) for v in df[col].tolist()]
            ws.column_dimensions[get_column_letter(i)].width = min(max(largos) + 2, 42)

        # Formato moneda en la columna Monto
        col_monto = get_column_letter(EXPORT_COLS.index("Monto") + 1)
        for r in range(2, nfilas + 2):
            ws[f"{col_monto}{r}"].number_format = '#,##0.00'

    return buf.getvalue()


# ----- Sidebar ----------------------------------------------------------------

st.sidebar.title("🧾 Facturas de Servicios")
seccion = st.sidebar.radio(
    "Sección",
    ["📊 Consultar", "📈 Dashboard", "📥 A descargar", "➕ Cargar factura", "💳 Pagos",
     "🏢 Proveedores", "🗑️ Borrar / corregir"],
)
if st.sidebar.button("🔄 Actualizar datos"):
    _refrescar()
    st.rerun()

# Chequeo de conexión temprano, con mensaje claro
try:
    facturas = cargar_facturas()
    proveedores = cargar_proveedores()
except Exception as e:  # noqa: BLE001
    st.error(
        "No me pude conectar al Google Sheet. Revisá las credenciales "
        "(st.secrets en la nube, o GOOGLE_SA_PATH + FACTURAS_SHEETS_ID en .env).\n\n"
        f"Detalle: {e}"
    )
    st.stop()

st.sidebar.caption(f"{len(facturas)} facturas en la base")


# =============================================================================
# CONSULTAR
# =============================================================================
if seccion == "📊 Consultar":
    st.header("📊 Consultar facturas")

    # ---- Segmentación / filtros (se aplican a todas las pestañas y a la descarga) ----
    facturas_f, hay_filtro = aplicar_filtros(facturas)

    tab_pend, tab_pag, tab_hist = st.tabs(
        ["⏳ Pendientes de pago", "✅ Pagadas", "📅 Histórico por período"]
    )

    pendientes = [f for f in facturas_f if f.get("estado_pago", db.ESTADO_PENDIENTE) != db.ESTADO_PAGADA]
    pagadas = [f for f in facturas_f if f.get("estado_pago") == db.ESTADO_PAGADA]

    # ---- Pendientes
    with tab_pend:
        if not pendientes:
            st.success("No hay facturas pendientes. 🎉")
        else:
            filas = []
            total = 0.0
            for f in sorted(pendientes, key=lambda x: fl.parse_fecha(x.get("primer_vto")) or date.max):
                cod, et, emoji = estado_venc(f)
                if f.get("monto_num"):
                    total += f["monto_num"]
                filas.append({
                    "Estado": f"{emoji} {et}",
                    "Proveedor": f.get("proveedor", ""),
                    "Cuenta": f.get("cuenta", ""),
                    "Nº Cliente": f.get("nro_cliente", ""),
                    "Vencimiento": f.get("primer_vto", ""),
                    "Monto": fmt_money(f.get("monto_num")),
                    "Rubro": f.get("rubro", ""),
                    "Período": f.get("periodo", ""),
                    "Comprobante": f.get("comprobante", ""),
                    "Factura": f.get("factura_url", ""),
                })
            vencidas = sum(1 for f in pendientes if estado_venc(f)[0] == "vencida")
            por_vencer = sum(1 for f in pendientes if estado_venc(f)[0] == "por_vencer")
            c1, c2, c3 = st.columns(3)
            c1.metric("Pendientes", len(pendientes))
            c2.metric("🔴 Vencidas", vencidas)
            c3.metric("🟡 Por vencer (≤10d)", por_vencer)
            st.caption(f"Total pendiente: **{fmt_money(total)}**")
            st.dataframe(
                pd.DataFrame(filas), use_container_width=True, hide_index=True,
                column_config={"Factura": st.column_config.LinkColumn(
                    "Factura", display_text="📄 ver")},
            )

    # ---- Pagadas
    with tab_pag:
        if not pagadas:
            st.info("Todavía no hay facturas marcadas como pagadas.")
        else:
            filas = []
            for f in sorted(pagadas, key=lambda x: fl.parse_fecha(x.get("fecha_pago")) or date.min, reverse=True):
                filas.append({
                    "Proveedor": f.get("proveedor", ""),
                    "Cuenta": f.get("cuenta", ""),
                    "Vencimiento": f.get("primer_vto", ""),
                    "Monto": fmt_money(f.get("monto_num")),
                    "Rubro": f.get("rubro", ""),
                    "Pagada el": f.get("fecha_pago", ""),
                    "Pagó": f.get("pagado_por", ""),
                    "Comprobante": f.get("comprobante", ""),
                    "Período": f.get("periodo", ""),
                    "Factura": f.get("factura_url", ""),
                    "Comp. pago": f.get("comprobante_pago_url", ""),
                })
            st.metric("Pagadas", len(pagadas))
            st.dataframe(
                pd.DataFrame(filas), use_container_width=True, hide_index=True,
                column_config={
                    "Factura": st.column_config.LinkColumn("Factura", display_text="📄 ver"),
                    "Comp. pago": st.column_config.LinkColumn("Comp. pago", display_text="🧾 ver"),
                },
            )

    # ---- Histórico por período
    with tab_hist:
        st.caption("Pagos agrupados por mes de vencimiento.")
        if not pagadas:
            st.info("No hay pagos registrados todavía.")
        else:
            rows = []
            for f in pagadas:
                vto = fl.parse_fecha(f.get("primer_vto"))
                periodo = f.get("periodo") or (vto.strftime("%Y-%m") if vto else "sin fecha")
                rows.append({"Período": periodo, "Monto": f.get("monto_num") or 0.0})
            dfp = pd.DataFrame(rows)
            resumen = dfp.groupby("Período").agg(
                Facturas=("Monto", "count"), Total=("Monto", "sum")
            ).reset_index().sort_values("Período", ascending=False)
            resumen["Total"] = resumen["Total"].apply(fmt_money)
            st.dataframe(resumen, use_container_width=True, hide_index=True)

    # Exportar (respeta el filtro/segmentación activo)
    st.divider()
    cuales = "filtradas" if hay_filtro else "todas"
    st.caption(f"Descarga **{len(facturas_f)}** facturas ({cuales}). El Excel viene como "
               "**Tabla**: ya trae filtros por columna y podés agregar *Segmentación de datos* "
               "(en Excel → pestaña **Insertar → Segmentación de datos**).")
    dc1, dc2 = st.columns([2, 1])
    dc1.download_button(
        "⬇️ Descargar Excel (.xlsx)",
        exportar_excel(facturas_f),
        file_name="facturas_servicios.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
    dc2.download_button(
        "⬇️ CSV",
        df_export(facturas_f).to_csv(index=False).encode("utf-8-sig"),
        file_name="facturas_servicios.csv", mime="text/csv",
    )


# =============================================================================
# DASHBOARD (gráficos interactivos)
# =============================================================================
elif seccion == "📈 Dashboard":
    import plotly.express as px

    st.header("📈 Dashboard")
    st.caption("Gráficos interactivos (pasá el mouse para ver valores). Respetan el "
               "filtro 🔎 de abajo. Tip: en cada gráfico podés hacer zoom y, con la "
               "cámara 📷 arriba a la derecha, descargarlo como imagen.")

    facturas_f, _ = aplicar_filtros(facturas)

    if not facturas_f:
        st.info("No hay datos para mostrar con el filtro actual. Probá 🧹 Limpiar filtros.")
    else:
        # ---- KPIs ----
        pend = [f for f in facturas_f if f.get("estado_pago") != db.ESTADO_PAGADA]
        pag = [f for f in facturas_f if f.get("estado_pago") == db.ESTADO_PAGADA]
        venc = [f for f in pend if estado_venc(f)[0] == "vencida"]
        tot_pend = sum(f.get("monto_num") or 0.0 for f in pend)
        tot_pag = sum(f.get("monto_num") or 0.0 for f in pag)
        tot_venc = sum(f.get("monto_num") or 0.0 for f in venc)
        k1, k2, k3, k4 = st.columns(4)
        k1.metric("💰 Pendiente", fmt_money(tot_pend))
        k2.metric("✅ Pagado", fmt_money(tot_pag))
        k3.metric("🔴 Vencido", fmt_money(tot_venc))
        k4.metric("Facturas vencidas", len(venc))

        # ---- DataFrame base ----
        rows = []
        for f in facturas_f:
            psort, plabel = periodo_par(f)
            rows.append({
                "Período": plabel,
                "_orden": psort,
                "Proveedor": f.get("proveedor", "") or "(sin proveedor)",
                "Rubro": f.get("rubro", "") or "(sin rubro)",
                "Estado": "Pagada" if f.get("estado_pago") == db.ESTADO_PAGADA else "Pendiente",
                "Monto": f.get("monto_num") or 0.0,
            })
        df = pd.DataFrame(rows)

        COLOR_ESTADO = {"Pendiente": "#E8A33D", "Pagada": "#3DA35D"}

        st.divider()

        # ---- Gasto por mes ----
        st.subheader("Gasto por período (mes)")
        gm = (df.groupby(["Período", "_orden"], as_index=False)["Monto"].sum()
                .sort_values("_orden"))
        fig_mes = px.bar(gm, x="Período", y="Monto", text_auto=".2s",
                         labels={"Monto": "Monto ($)"})
        fig_mes.update_traces(marker_color="#4C78A8",
                              hovertemplate="Período %{x}<br>$ %{y:,.2f}<extra></extra>")
        fig_mes.update_layout(yaxis_tickprefix="$ ", showlegend=False, height=340,
                              margin=dict(t=10, b=0, l=0, r=0))
        st.plotly_chart(fig_mes, use_container_width=True)

        # ---- Rubro (torta) + Proveedor (barras) ----
        c1, c2 = st.columns(2)
        with c1:
            st.subheader("Gasto por rubro")
            gr = df.groupby("Rubro", as_index=False)["Monto"].sum().sort_values("Monto", ascending=False)
            fig_rub = px.pie(gr, names="Rubro", values="Monto", hole=0.45)
            fig_rub.update_traces(textposition="inside", texttemplate="%{label}<br>%{percent}",
                                  hovertemplate="%{label}<br>$ %{value:,.2f}<extra></extra>")
            fig_rub.update_layout(height=360, margin=dict(t=10, b=0, l=0, r=0))
            st.plotly_chart(fig_rub, use_container_width=True)
        with c2:
            st.subheader("Gasto por proveedor")
            gp = df.groupby("Proveedor", as_index=False)["Monto"].sum().sort_values("Monto")
            fig_prov = px.bar(gp, x="Monto", y="Proveedor", orientation="h", text_auto=".2s",
                              labels={"Monto": "Monto ($)"})
            fig_prov.update_traces(marker_color="#72B7B2",
                                   hovertemplate="%{y}<br>$ %{x:,.2f}<extra></extra>")
            fig_prov.update_layout(xaxis_tickprefix="$ ", height=360,
                                   margin=dict(t=10, b=0, l=0, r=0))
            st.plotly_chart(fig_prov, use_container_width=True)

        # ---- Pendiente vs Pagado por mes ----
        st.subheader("Pendiente vs Pagado por período")
        gpp = (df.groupby(["Período", "_orden", "Estado"], as_index=False)["Monto"].sum()
                 .sort_values("_orden"))
        fig_pp = px.bar(gpp, x="Período", y="Monto", color="Estado", barmode="group",
                        color_discrete_map=COLOR_ESTADO, labels={"Monto": "Monto ($)"})
        fig_pp.update_traces(hovertemplate="%{x}<br>%{fullData.name}: $ %{y:,.2f}<extra></extra>")
        fig_pp.update_layout(yaxis_tickprefix="$ ", height=360,
                             margin=dict(t=10, b=0, l=0, r=0),
                             legend=dict(orientation="h", y=1.1))
        st.plotly_chart(fig_pp, use_container_width=True)


# =============================================================================
# A DESCARGAR (ayuda memoria de emisión)
# =============================================================================
elif seccion == "📥 A descargar":
    st.header("📥 Ayuda memoria — facturas a buscar/descargar")
    st.caption("Según la fecha típica de emisión de cada cuenta, te dice cuáles "
               "**ya deberían tener una factura nueva emitida** (para ir a descargarla del portal).")

    ref = st.date_input("Fecha de consulta", value=date.today(), format="DD/MM/YYYY")
    nombre_dl = st.text_input("Tu nombre (opcional, queda registrado al marcar)", key="dl_nombre")

    prep = [fl.preparar_factura(f) for f in facturas]
    cuentas = fl.agrupar_cuentas(prep)
    marcadas = cargar_descargas()

    def _clave_dl(c, r):
        return (f"{c['proveedor']}|{c['cuenta']}|{c['nro_cliente']}|"
                f"{fl.fmt_fecha(r['emision_esperada'])}")

    disponibles, proximas = [], []
    for c in cuentas:
        r = fl.ayuda_memoria_descarga(c, ref)
        if r["estado"] == "disponible":
            disponibles.append((c, r))
        elif r["estado"] == "proxima":
            proximas.append((c, r))

    disponibles.sort(key=lambda x: x[1]["emision_esperada"])
    proximas.sort(key=lambda x: x[1]["emision_esperada"])

    pendientes = [(c, r) for c, r in disponibles if _clave_dl(c, r) not in marcadas]
    ya_desc = [(c, r) for c, r in disponibles if _clave_dl(c, r) in marcadas]

    st.subheader(f"🟢 Ya deberían estar emitidas ({len(pendientes)}) — buscalas en el portal")
    if not pendientes:
        st.success("No hay facturas nuevas para descargar por ahora. 🎉")
    else:
        for c, r in pendientes:
            clave = _clave_dl(c, r)
            with st.container(border=True):
                cols = st.columns([4, 2, 3, 2])
                cols[0].markdown(f"**{c['proveedor']}** · {c['cuenta']}")
                cols[1].markdown(f"Nº {c['nro_cliente'] or '—'}")
                cols[2].markdown(
                    f"Emisión esperada: **{fl.fmt_fecha(r['emision_esperada'])}**  \n"
                    f"({'estimada' if r['estimada'] else 'real'})"
                )
                if cols[3].button("✅ Ya la descargué", key=f"dl_{clave}", type="primary"):
                    db.marcar_descarga(clave, c["proveedor"], c["cuenta"], c["nro_cliente"],
                                       fl.fmt_fecha(r["emision_esperada"]), nombre_dl.strip())
                    _refrescar()
                    st.rerun()

    if ya_desc:
        with st.expander(f"✓ Ya descargadas este período ({len(ya_desc)})"):
            for c, r in ya_desc:
                clave = _clave_dl(c, r)
                cc = st.columns([6, 2])
                cc[0].markdown(f"{c['proveedor']} · {c['cuenta']} — emisión "
                               f"{fl.fmt_fecha(r['emision_esperada'])}")
                if cc[1].button("↩️ Deshacer", key=f"undl_{clave}"):
                    db.desmarcar_descarga(clave)
                    _refrescar()
                    st.rerun()
            st.caption("Cuando salga la factura del mes siguiente, la cuenta vuelve a "
                       "aparecer sola arriba.")

    st.divider()
    st.subheader(f"🔜 Próximas a emitir ({len(proximas)})")
    if proximas:
        filas = []
        for c, r in proximas:
            filas.append({
                "Proveedor": c["proveedor"],
                "Cuenta": c["cuenta"],
                "Nº Cliente": c["nro_cliente"],
                "Próx. emisión ~": fl.fmt_fecha(r["emision_esperada"]),
                "Base": "estimada" if r["estimada"] else "real",
            })
        st.dataframe(pd.DataFrame(filas), use_container_width=True, hide_index=True)

    st.caption("ℹ️ Las cuentas marcadas como **estimada** (AySA, Aguas Cordobesas, Epec, etc.) "
               "no informan fecha de emisión en el portal: se estima como vencimiento − 12 días. "
               "Es orientativo.")


# =============================================================================
# CARGAR FACTURA
# =============================================================================
elif seccion == "➕ Cargar factura":
    st.header("➕ Cargar / editar factura")

    if "cf_ok_msg" in st.session_state:
        st.success(st.session_state.pop("cf_ok_msg"))
    if "cf_edit_ok_msg" in st.session_state:
        st.success(st.session_state.pop("cf_edit_ok_msg"))

    # ---- Buscar una factura ya cargada para editarla (por N° de comprobante) ----
    with st.expander("🔍 Buscar una factura ya cargada (por N° de comprobante) para editarla",
                     expanded=bool(st.session_state.get("cf_edit_fid"))):
        cb1, cb2 = st.columns([4, 1])
        q_comp = cb1.text_input("N° de comprobante (o parte)", key="cf_busca_comp")
        if cb2.button("🔍 Buscar", key="cf_busca_btn"):
            qn = (q_comp or "").strip()
            qn_fmt = formatear_comprobante(qn)
            ql = qn.lower()
            matches = []
            if qn:
                for f in facturas:
                    comp = str(f.get("comprobante", ""))
                    if (qn_fmt and comp == qn_fmt) or (ql and ql in comp.lower()):
                        matches.append(f)
            st.session_state["cf_edit_matches"] = [m["id"] for m in matches]
            st.session_state["cf_edit_fid"] = matches[0]["id"] if len(matches) == 1 else None
            st.session_state["cf_busca_hecha"] = True

        match_ids = st.session_state.get("cf_edit_matches", [])
        cand = [f for f in facturas if f["id"] in match_ids]
        if len(cand) > 1:
            op = {"— Elegí la factura —": None}
            for f in cand:
                op[(f"{f.get('proveedor','')} · {f.get('cuenta','')} · vto "
                    f"{f.get('primer_vto','')} · {fmt_money(f.get('monto_num'))} · "
                    f"comp {f.get('comprobante','—')}  [{f['id']}]")] = f["id"]
            sel = st.selectbox("Resultados", list(op.keys()), key="cf_edit_sel")
            if op[sel]:
                st.session_state["cf_edit_fid"] = op[sel]
        elif not cand and st.session_state.get("cf_busca_hecha"):
            st.info("Sin resultados para ese comprobante.")

    edit_fid = st.session_state.get("cf_edit_fid")
    fe = next((x for x in facturas if x["id"] == edit_fid), None) if edit_fid else None

    # =========================================================================
    # MODO EDICIÓN — corregir una factura existente (incluye asignar rubro)
    # =========================================================================
    if fe:
        st.subheader(f"✏️ Editando: {fe.get('proveedor','')} · {fe.get('cuenta','')} "
                     f"· comp {fe.get('comprobante','—')}")
        st.caption(f"id {fe['id']} — se modifica en el lugar (no borra ni duplica).")
        k = f"ed_{fe['id']}"
        e1, e2 = st.columns(2)
        with e1:
            ed_prov = st.text_input("Proveedor", value=fe.get("proveedor", ""), key=f"{k}_prov")
            ed_cuenta = st.text_input("Cuenta", value=fe.get("cuenta", ""), key=f"{k}_cuenta")
            ed_nrocli = st.text_input("Nº Cliente", value=fe.get("nro_cliente", ""), key=f"{k}_nro")
            ed_emision = st.date_input("Fecha de emisión",
                                       value=fl.parse_fecha(fe.get("fecha_emision")),
                                       format="DD/MM/YYYY", key=f"{k}_emision")
            ed_vto = st.date_input("1er vencimiento", value=fl.parse_fecha(fe.get("primer_vto")),
                                   format="DD/MM/YYYY", key=f"{k}_vto")
        with e2:
            ed_vto2 = st.date_input("2do vencimiento (opcional)",
                                    value=fl.parse_fecha(fe.get("segundo_vto")),
                                    format="DD/MM/YYYY", key=f"{k}_vto2")
            ed_monto = st.text_input("Monto", value=fmt_monto_ar_num(fe.get("monto_num")),
                                     key=f"{k}_monto")
            ed_comp = st.text_input("N° de comprobante", value=fe.get("comprobante", ""),
                                    key=f"{k}_comp")
            ed_periodo = st.text_input("Período", value=fe.get("periodo", ""), key=f"{k}_periodo")
        ed_rubro = _input_rubro(k, fe.get("rubro", ""))
        ed_nota = st.text_input("Nota", value=fe.get("nota", ""), key=f"{k}_nota")

        bcols = st.columns([1, 1, 4])
        if bcols[0].button("💾 Guardar cambios", type="primary", key=f"{k}_save"):
            monto_val = parse_monto_ar(ed_monto)
            comp_fmt = formatear_comprobante(ed_comp) if ed_comp.strip() else ""
            cambios = {
                "proveedor": ed_prov.strip(),
                "cuenta": ed_cuenta.strip(),
                "nro_cliente": ed_nrocli.strip(),
                "fecha_emision": fl.fmt_fecha(ed_emision) if ed_emision else "",
                "primer_vto": fl.fmt_fecha(ed_vto) if ed_vto else "",
                "segundo_vto": fl.fmt_fecha(ed_vto2) if ed_vto2 else "",
                "monto": monto_val if monto_val is not None else "",
                "periodo": ed_periodo.strip(),
                "comprobante": comp_fmt or ed_comp.strip(),
                "rubro": ed_rubro,
                "nota": ed_nota.strip(),
            }
            db.actualizar_factura(fe["id"], cambios)
            _refrescar()
            for kk in ("cf_edit_fid", "cf_edit_matches", "cf_busca_hecha"):
                st.session_state.pop(kk, None)
            st.session_state["cf_edit_ok_msg"] = (
                f"Cambios guardados ✅ — {ed_prov.strip()} · {ed_cuenta.strip()} "
                f"(rubro: {ed_rubro or '—'})"
            )
            st.rerun()
        if bcols[1].button("✖️ Cancelar / cargar nueva", key=f"{k}_cancel"):
            for kk in ("cf_edit_fid", "cf_edit_matches", "cf_busca_hecha"):
                st.session_state.pop(kk, None)
            st.rerun()

    # =========================================================================
    # MODO NUEVA — alta de factura
    # =========================================================================
    else:
        # Limpiar campos de la carga anterior ANTES de instanciar los widgets
        if st.session_state.pop("cf_clear", False):
            for kk in ("cf_emision", "cf_vto", "cf_vto2", "cf_monto", "cf_comp", "cf_nota",
                       "cf_factura_file", "cf_rubro", "cf_rubro_otro"):
                st.session_state.pop(kk, None)

        nombres_prov = sorted({p["proveedor"] for p in proveedores if p.get("proveedor")})
        opciones = ["— Elegí proveedor —"] + nombres_prov + ["➕ Nuevo proveedor (cargar en 🏢 Proveedores)"]

        sel_prov = st.selectbox("Proveedor", opciones, key="cf_prov")

        # Si cambió el proveedor, resetear la cuenta elegida (antes de instanciar el selectbox)
        if st.session_state.get("cf_prov_prev") != sel_prov:
            st.session_state.pop("cf_cuenta_idx", None)
            st.session_state["cf_prov_prev"] = sel_prov

        proveedor = ""
        cuenta_val = ""
        nrocli_val = ""
        if sel_prov in nombres_prov:
            proveedor = sel_prov
            cuentas_prov = [p for p in proveedores if p.get("proveedor") == sel_prov]

            def _etiqueta_cuenta(p):
                cta = p.get("cuenta", "")
                nro = p.get("nro_cliente", "")
                if nro and nro not in cta:
                    return f"{cta} · {nro}"
                return cta

            etiquetas = [_etiqueta_cuenta(p) for p in cuentas_prov]
            # Primera opción = placeholder vacío (-1): obliga a elegir la cuenta.
            idx = st.selectbox(
                "Cuenta",
                [-1] + list(range(len(cuentas_prov))),
                format_func=lambda i: "— Elegí la cuenta —" if i == -1 else etiquetas[i],
                key="cf_cuenta_idx",
            )
            if idx is not None and idx >= 0:
                cuenta_val = cuentas_prov[idx].get("cuenta", "")
                nrocli_val = cuentas_prov[idx].get("nro_cliente", "")
        elif sel_prov.startswith("➕"):
            st.info("Para un proveedor nuevo, primero cargalo en la sección **🏢 Proveedores** y volvé acá.")

        # --- Campos (sin formulario: reaccionan al instante) ---
        c1, c2 = st.columns(2)
        with c1:
            # Bloqueados: se autocompletan según la cuenta elegida
            st.text_input("Cuenta", value=cuenta_val, disabled=True)
            st.text_input("Nº Cliente", value=nrocli_val, disabled=True)
            emision = st.date_input("Fecha de emisión *", value=None, format="DD/MM/YYYY", key="cf_emision")
            primer_vto = st.date_input("1er vencimiento *", value=None, format="DD/MM/YYYY", key="cf_vto")
        with c2:
            segundo_vto = st.date_input("2do vencimiento (opcional)", value=None,
                                        format="DD/MM/YYYY", key="cf_vto2")
            monto_str = st.text_input("Monto *", placeholder="27.000,00", key="cf_monto",
                                      on_change=_autoformatear_monto,
                                      help="Formato argentino: '.' para miles y ',' para decimales.")
            comprobante = st.text_input("N° de comprobante *", placeholder="67 - 8878", key="cf_comp",
                                        on_change=_autoformatear_comp,
                                        help="Se guarda como 00067 - 00008878 (5 + 8 dígitos).")
            # Período: automático y bloqueado
            periodo_val = periodo_auto(primer_vto) if primer_vto else ""
            st.text_input("Período (automático)", value=periodo_val, disabled=True,
                          help="Mes anterior al 1er vencimiento. Se completa solo.")
        rubro_val = _input_rubro("cf")
        nota = st.text_input("Nota", key="cf_nota")
        archivo_factura = st.file_uploader(
            "📎 Adjuntar factura (PDF o imagen) — opcional",
            type=["pdf", "jpg", "jpeg", "png"], key="cf_factura_file",
        )

        if st.button("💾 Guardar factura", type="primary"):
            monto_val = parse_monto_ar(monto_str)
            comp_fmt = formatear_comprobante(comprobante) if comprobante.strip() else None
            faltan = []
            if not proveedor:
                faltan.append("Proveedor")
            if not cuenta_val:
                faltan.append("Cuenta")
            if emision is None:
                faltan.append("Fecha de emisión")
            if primer_vto is None:
                faltan.append("1er vencimiento")
            if monto_val is None or monto_val <= 0:
                faltan.append("Monto")
            if not comprobante.strip():
                faltan.append("N° de comprobante")

            if faltan:
                st.error("Faltan campos obligatorios: " + ", ".join(faltan))
            elif comp_fmt is None:
                st.error("El N° de comprobante debe tener dos números, ej: **67 - 8878** "
                         "(se guarda como 00057 - 00089898).")
            else:
                periodo = periodo_auto(primer_vto)
                factura_url = ""
                if archivo_factura is not None:
                    try:
                        with st.spinner("Subiendo la factura a Drive…"):
                            factura_url = drive.subir_comprobante(
                                archivo_factura.getvalue(), archivo_factura.name,
                                archivo_factura.type,
                                prefijo=f"{proveedor}_{cuenta_val}_{periodo}_factura",
                            )
                    except Exception as e:  # noqa: BLE001
                        st.warning("No pude subir el adjunto a Drive (la factura se guarda "
                                   f"igual; podés adjuntarlo después). Detalle: {e}")
                fid = db.append_factura({
                    "proveedor": proveedor,
                    "cuenta": cuenta_val,
                    "nro_cliente": nrocli_val,
                    "fecha_emision": fl.fmt_fecha(emision),
                    "primer_vto": fl.fmt_fecha(primer_vto),
                    "segundo_vto": fl.fmt_fecha(segundo_vto) if segundo_vto else "",
                    "monto": monto_val,
                    "periodo": periodo,
                    "comprobante": comp_fmt,
                    "nota": nota,
                    "origen": "manual",
                    "estado_pago": db.ESTADO_PENDIENTE,
                    "factura_url": factura_url,
                    "rubro": rubro_val,
                })
                _refrescar()
                # Bandera para limpiar los campos en el próximo run (mantiene proveedor/cuenta)
                st.session_state["cf_clear"] = True
                st.session_state["cf_ok_msg"] = (
                    f"Factura guardada ✅ — comp. **{comp_fmt}**, período **{periodo}**, "
                    f"monto **{fmt_money(monto_val)}**, rubro **{rubro_val or '—'}** (id {fid})"
                )
                st.rerun()


# =============================================================================
# PAGOS
# =============================================================================
elif seccion == "💳 Pagos":
    st.header("💳 Pagos — buscar y marcar pagado")

    c1, c2 = st.columns([3, 1])
    busqueda = c1.text_input("Buscar (proveedor, cuenta, nº cliente, comprobante o período)")
    solo_pend = c2.checkbox("Solo pendientes", value=True)
    pagado_por = st.text_input("Tu nombre (queda registrado en 'Pagó')", key="pagado_por")

    def coincide(f):
        if solo_pend and f.get("estado_pago") == db.ESTADO_PAGADA:
            return False
        if not busqueda:
            return True
        q = busqueda.lower()
        campos = [f.get(k, "") for k in
                  ("proveedor", "cuenta", "nro_cliente", "comprobante", "periodo")]
        return any(q in str(v).lower() for v in campos)

    resultados = [f for f in facturas if coincide(f)]
    resultados.sort(key=lambda x: fl.parse_fecha(x.get("primer_vto")) or date.max)

    st.caption(f"{len(resultados)} factura(s)")
    if not resultados:
        st.info("Sin resultados. Ajustá la búsqueda.")

    for f in resultados[:200]:
        cod, et, emoji = estado_venc(f)
        pagada = f.get("estado_pago") == db.ESTADO_PAGADA
        with st.container(border=True):
            cols = st.columns([4, 2, 2, 2])
            cols[0].markdown(
                f"**{f.get('proveedor','')}** · {f.get('cuenta','')}  \n"
                f"Cliente {f.get('nro_cliente','—')} · Comp. {f.get('comprobante','—')}"
            )
            cols[1].markdown(f"Vto: **{f.get('primer_vto','')}**  \n{emoji} {et}")
            cols[2].markdown(f"Monto:  \n**{fmt_money(f.get('monto_num'))}**")
            with cols[3]:
                fact_url = f.get("factura_url", "")
                pago_url = f.get("comprobante_pago_url", "")
                if fact_url:
                    st.markdown(f"[📄 Factura]({fact_url})")
                if pago_url:
                    st.markdown(f"[🧾 Comprobante de pago]({pago_url})")

                # Adjuntar la factura (PDF) a un registro ya cargado que no la tiene.
                if not fact_url:
                    fac_file = st.file_uploader(
                        "Adjuntar la factura (PDF/imagen)",
                        type=["pdf", "jpg", "jpeg", "png"], key=f"fac_file_{f['id']}",
                    )
                    if fac_file is not None and st.button(
                            "⬆️ Subir factura", key=f"upfac_{f['id']}"):
                        try:
                            with st.spinner("Subiendo la factura a Drive…"):
                                url = drive.subir_comprobante(
                                    fac_file.getvalue(), fac_file.name, fac_file.type,
                                    prefijo=f"{f.get('proveedor','')}_{f.get('cuenta','')}_"
                                           f"{f.get('periodo','')}_factura",
                                )
                            db.set_url_adjunto(f["id"], db.COL_FACTURA_URL, url)
                            _refrescar()
                            st.rerun()
                        except Exception as e:  # noqa: BLE001
                            st.warning(f"No pude subir el adjunto. Detalle: {e}")

                if pagada:
                    st.success(f"✅ Pagada {f.get('fecha_pago','')}")
                    if not pago_url:
                        comp_pago = st.file_uploader(
                            "Adjuntar comprobante de pago",
                            type=["pdf", "jpg", "jpeg", "png"], key=f"pago_file_{f['id']}",
                        )
                        if comp_pago is not None and st.button(
                                "⬆️ Subir comprobante", key=f"uppago_{f['id']}"):
                            try:
                                with st.spinner("Subiendo comprobante…"):
                                    url = drive.subir_comprobante(
                                        comp_pago.getvalue(), comp_pago.name, comp_pago.type,
                                        prefijo=f"{f.get('proveedor','')}_{f.get('cuenta','')}_pago",
                                    )
                                db.set_url_adjunto(f["id"], db.COL_COMPROBANTE_PAGO_URL, url)
                                _refrescar()
                                st.rerun()
                            except Exception as e:  # noqa: BLE001
                                st.warning(f"No pude subir el adjunto. Detalle: {e}")
                    if st.button("↩️ Revertir", key=f"rev_{f['id']}"):
                        db.marcar_pendiente(f["id"])
                        _refrescar()
                        st.rerun()
                else:
                    comp_pago = st.file_uploader(
                        "Comprobante de pago — opcional",
                        type=["pdf", "jpg", "jpeg", "png"], key=f"pago_file_{f['id']}",
                    )
                    if st.button("✔️ Marcar pagada", key=f"pay_{f['id']}", type="primary"):
                        url = ""
                        if comp_pago is not None:
                            try:
                                with st.spinner("Subiendo comprobante…"):
                                    url = drive.subir_comprobante(
                                        comp_pago.getvalue(), comp_pago.name, comp_pago.type,
                                        prefijo=f"{f.get('proveedor','')}_{f.get('cuenta','')}_pago",
                                    )
                            except Exception as e:  # noqa: BLE001
                                st.warning("No pude subir el adjunto (igual marco pagada). "
                                           f"Detalle: {e}")
                        db.marcar_pagada(f["id"], pagado_por=pagado_por.strip())
                        if url:
                            db.set_url_adjunto(f["id"], db.COL_COMPROBANTE_PAGO_URL, url)
                        _refrescar()
                        st.toast(f"Pagada: {f.get('proveedor','')} {f.get('cuenta','')}")
                        st.rerun()


# =============================================================================
# PROVEEDORES
# =============================================================================
elif seccion == "🏢 Proveedores":
    st.header("🏢 Proveedores")
    st.caption("Acá agregás proveedores/cuentas para que aparezcan en el formulario de carga.")

    if proveedores:
        st.dataframe(
            pd.DataFrame([{k: p.get(k, "") for k in db.PROVEEDORES_HEADERS} for p in proveedores]),
            use_container_width=True, hide_index=True,
        )
    else:
        st.info("Todavía no hay proveedores cargados.")

    st.subheader("➕ Agregar proveedor / cuenta")
    with st.form("alta_prov", clear_on_submit=True):
        c1, c2 = st.columns(2)
        proveedor = c1.text_input("Proveedor *")
        cuenta = c2.text_input("Cuenta *")
        nro_cliente = c1.text_input("Nº Cliente")
        trae_emision = c2.selectbox("¿El portal informa fecha de emisión?", ["No", "Sí"])
        nota = st.text_input("Nota")
        if st.form_submit_button("💾 Guardar proveedor", type="primary"):
            if not proveedor or not cuenta:
                st.error("Proveedor y Cuenta son obligatorios.")
            else:
                db.append_proveedor(proveedor.strip(), cuenta.strip(),
                                    nro_cliente.strip(), trae_emision, nota.strip())
                _refrescar()
                st.success(f"Proveedor agregado ✅ ({proveedor} · {cuenta})")
                st.rerun()


# =============================================================================
# BORRAR / CORREGIR
# =============================================================================
elif seccion == "🗑️ Borrar / corregir":
    st.header("🗑️ Borrar registros")
    st.info("Para **corregir** un dato no se edita: se **elimina** el registro y se "
            "vuelve a cargar correcto (la factura en **➕ Cargar factura**, el proveedor "
            "en **🏢 Proveedores**). Borrar pide la clave del equipo.")

    if "del_ok_msg" in st.session_state:
        st.success(st.session_state.pop("del_ok_msg"))

    tab_f, tab_p = st.tabs(["Facturas", "Proveedores"])

    # ---- Borrar factura
    with tab_f:
        q = st.text_input(
            "Buscar la factura (proveedor, cuenta, nº cliente, comprobante o período)",
            key="del_f_q",
        )
        cand = facturas
        if q:
            ql = q.lower()
            cand = [f for f in facturas if any(
                ql in str(f.get(k, "")).lower()
                for k in ("proveedor", "cuenta", "nro_cliente", "comprobante", "periodo"))]
        cand = sorted(cand, key=lambda x: fl.parse_fecha(x.get("primer_vto")) or date.max)

        if not cand:
            st.info("Sin resultados para esa búsqueda.")
        else:
            def _lbl_f(f):
                return (f"{f.get('proveedor','')} · {f.get('cuenta','')} · "
                        f"vto {f.get('primer_vto','')} · {fmt_money(f.get('monto_num'))} · "
                        f"comp {f.get('comprobante','—')}  [{f['id']}]")

            opciones = {"— Elegí una factura —": None}
            for f in cand[:200]:
                opciones[_lbl_f(f)] = f["id"]
            sel = st.selectbox("Factura a eliminar", list(opciones.keys()), key="del_f_sel")
            fid = opciones[sel]
            if fid:
                f = next((x for x in facturas if x["id"] == fid), None)
                st.warning(
                    f"Vas a **ELIMINAR** esta factura:\n\n"
                    f"- **{f.get('proveedor','')} · {f.get('cuenta','')}**\n"
                    f"- Vencimiento: {f.get('primer_vto','')} · Monto: {fmt_money(f.get('monto_num'))}\n"
                    f"- Comprobante: {f.get('comprobante','—')} · Período: {f.get('periodo','—')}"
                )
                clave_f = st.text_input("Clave del equipo para confirmar", type="password",
                                        key="del_f_pwd")
                if st.button("🗑️ Eliminar factura", type="primary", key="del_f_btn"):
                    if clave_f != _clave_app():
                        st.error("Clave incorrecta. No se borró nada.")
                    else:
                        db.borrar_factura(fid)
                        _refrescar()
                        st.session_state["del_ok_msg"] = (
                            "Factura eliminada. Si fue un error de carga, volvé a cargarla "
                            "correcta en ➕ Cargar factura."
                        )
                        st.rerun()

    # ---- Borrar proveedor
    with tab_p:
        if not proveedores:
            st.info("No hay proveedores cargados.")
        else:
            opciones = {"— Elegí un proveedor —": None}
            for p in proveedores:
                lbl = f"{p.get('proveedor','')} · {p.get('cuenta','')}"
                if p.get("nro_cliente"):
                    lbl += f" · {p.get('nro_cliente')}"
                opciones[lbl] = p.get("_row")
            sel = st.selectbox("Proveedor a eliminar", list(opciones.keys()), key="del_p_sel")
            row = opciones[sel]
            if row:
                st.warning(f"Vas a **ELIMINAR** el proveedor/cuenta: **{sel}**")
                st.caption("No borra las facturas ya cargadas de ese proveedor, sólo la "
                           "cuenta del listado de carga.")
                clave_p = st.text_input("Clave del equipo para confirmar", type="password",
                                        key="del_p_pwd")
                if st.button("🗑️ Eliminar proveedor", type="primary", key="del_p_btn"):
                    if clave_p != _clave_app():
                        st.error("Clave incorrecta. No se borró nada.")
                    else:
                        db.borrar_proveedor(row)
                        _refrescar()
                        st.session_state["del_ok_msg"] = (
                            "Proveedor eliminado. Si fue un error, volvé a cargarlo en 🏢 Proveedores."
                        )
                        st.rerun()
